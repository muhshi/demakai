"""
Baca evaluasi_after_fix_21apr.xlsx untuk mendapatkan per-query data
Hybrid+Expansion (metode final) per kolom Hybrid_Expansion.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import load_workbook
    wb = load_workbook(r'd:\magang_bps\backend-demakai\python\output\evaluasi_after_fix_21apr.xlsx')
    ws = wb.active
    print(f"Sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()

    # Print header rows (1-4)
    for r in range(1, 5):
        row_vals = []
        for c in range(1, min(ws.max_column+1, 35)):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(f"[{c}]{v}")
        if row_vals:
            print(f"Row {r}: {' | '.join(row_vals[:12])}")
    print()

    # Cari kolom dengan label Hybrid_Expansion
    header_row_1 = {ws.cell(1, c).value: c for c in range(1, ws.max_column+1) if ws.cell(1,c).value}
    header_row_2 = {ws.cell(2, c).value: c for c in range(1, ws.max_column+1) if ws.cell(2,c).value}
    print("Row1 headers:", dict(list(header_row_1.items())[:15]))
    print("Row2 headers:", dict(list(header_row_2.items())[:15]))

except Exception as e:
    print(f"Error: {e}")
    import traceback; traceback.print_exc()
