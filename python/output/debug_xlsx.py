"""
Debug xlsx row format to find correct data rows.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'd:\magang_bps\backend-demakai\python\output\evaluasi_after_fix_21apr.xlsx')
ws = wb.active
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Print rows 1-10, columns 1-6 and 33-38
print("\nRows 1-10, cols 1-6 + cols 31-38:")
for r in range(1, 11):
    vals_left  = [repr(ws.cell(r, c).value) for c in range(1, 7)]
    vals_right = [repr(ws.cell(r, c).value) for c in range(31, 39)]
    print(f"Row {r:2d}: LEFT={vals_left}  RIGHT={vals_right}")

print("\nRows 35-45 (around KBJI separator):")
for r in range(35, 47):
    vals = [repr(ws.cell(r, c).value) for c in range(1, 7)]
    print(f"Row {r:2d}: {vals}")
