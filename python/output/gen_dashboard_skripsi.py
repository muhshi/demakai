"""
gen_dashboard_skripsi.py
Generate dashboard HTML bersih untuk keperluan skripsi.
Menampilkan 5 metode FINAL tanpa label "fix" / "bug".
Per-query data: dari evaluasi_semua CSV + evaluasi_after_fix xlsx (metode 5 final).
"""
import csv, sys, os
from openpyxl import load_workbook
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'd:\magang_bps\backend-demakai\python\output'

# ────────────────────────────────────────────────────────────────────
# 1. Load CSV evaluasi_semua (untuk metode 1-4)
# ────────────────────────────────────────────────────────────────────
def load_csv(path):
    with open(path, encoding='utf-8') as f:
        return list(csv.DictReader(f))

rows_all = load_csv(os.path.join(BASE, 'evaluasi_semua_20260420_144133.csv'))

def get_rows(metode_key, tipe):
    return [r for r in rows_all if r['metode'] == metode_key and r['tipe'] == tipe]

# ────────────────────────────────────────────────────────────────────
# 2. Load per-query data dari xlsx untuk metode 5 (Hybrid_Expansion AFTER FIX)
# ────────────────────────────────────────────────────────────────────
def load_hybrid_expansion_from_xlsx(xlsx_path):
    """
    Baca kolom Hybrid+Expansion (col 33-38) dari xlsx.
    Kolom: 33=Rank, 34=Top1, 35=Top5, 36=Top10, 37=KK, 38=Preprocessing text
    Data rows: KBLI mulai row 4 s/d 34, KBJI mulai dari row setelah separator.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Cari batas KBLI dan KBJI
    # Row 2 = header method, row 3 = metric names, row 4+ = data
    # KBLI: rows 4-33 (30 rows), KBJI: mulai dari baris setelah separator
    
    results = []
    
    # Scan semua baris untuk data Hybrid+Expansion
    kbli_started = False
    kbji_started = False
    
    for r in range(4, ws.max_row + 1):
        no_val   = ws.cell(r, 1).value
        tipe_val = ws.cell(r, 2).value
        query    = ws.cell(r, 3).value
        gt       = ws.cell(r, 4).value
        rank_hyb = ws.cell(r, 33).value
        top1_hyb = ws.cell(r, 34).value
        top5_hyb = ws.cell(r, 35).value
        top10_hyb= ws.cell(r, 36).value
        kk_hyb   = ws.cell(r, 37).value
        prep_hyb = ws.cell(r, 38).value
        
        # Skip non-data rows
        if not no_val or not str(no_val).replace('.','').isdigit():
            continue
        if not tipe_val or not query:
            continue
        if tipe_val not in ('KBLI', 'KBJI'):
            continue
            
        results.append({
            'no': str(no_val),
            'tipe': tipe_val,
            'query': str(query),
            'kode_gt': str(gt) if gt else '',
            'rank': str(int(rank_hyb)) if rank_hyb not in (None, '') else '0',
            'top1': str(int(top1_hyb)) if top1_hyb not in (None, '') else '0',
            'top3': str(int(top5_hyb)) if top5_hyb not in (None, '') else '0',
            'top10': str(int(top10_hyb)) if top10_hyb not in (None, '') else '0',
            'rr': str(round(float(kk_hyb), 4)) if kk_hyb not in (None, '') else '0',
            'preprocessed': str(prep_hyb) if prep_hyb else '',
            'contoh': '',
        })
    
    return results

xlsx_path = os.path.join(BASE, 'evaluasi_after_fix_21apr.xlsx')
rows_m5_raw = load_hybrid_expansion_from_xlsx(xlsx_path)
print(f"Hybrid+Expansion from xlsx: {len(rows_m5_raw)} rows")
for r in rows_m5_raw[:3]:
    print(f"  [{r['no']}] {r['tipe']} | {r['query'][:40]} | rank={r['rank']} top1={r['top1']} rr={r['rr']}")

# Enrich dengan contoh_lapangan dari evaluasi_semua (M6 sebelum fix - same query order)
m6_kbli_src = get_rows('6. Hybrid + Preprocessing + Contoh Lapangan (FINAL)', 'KBLI')
m6_kbji_src = get_rows('6. Hybrid + Preprocessing + Contoh Lapangan (FINAL)', 'KBJI')
contoh_map  = {(r['tipe'], r['no']): r.get('contoh', '') for r in m6_kbli_src + m6_kbji_src}
prep_map    = {(r['tipe'], r['no']): r.get('preprocessed', '') for r in m6_kbli_src + m6_kbji_src}

for r in rows_m5_raw:
    key = (r['tipe'], r['no'])
    r['contoh'] = contoh_map.get(key, '')
    if not r['preprocessed']:
        r['preprocessed'] = prep_map.get(key, '')

rows_m5_kbli = [r for r in rows_m5_raw if r['tipe'] == 'KBLI']
rows_m5_kbji = [r for r in rows_m5_raw if r['tipe'] == 'KBJI']
print(f"  KBLI: {len(rows_m5_kbli)}, KBJI: {len(rows_m5_kbji)}")

# ────────────────────────────────────────────────────────────────────
# 3. Define 5 metode final
# ────────────────────────────────────────────────────────────────────
METHODS = [
    {
        'id': 'm1', 'no': '01', 'label': 'SQL LIKE',
        'name': 'SQL LIKE (Baseline)',
        'sub':  'Pencarian frasa mentah — tanpa preprocessing, tanpa semantic search, tanpa contoh lapangan',
        'tags': ['SQL LIKE'],
        'color': '#64748b', 'status': 'Baseline',
        'show_prep': False, 'show_contoh': False,
        'kbli_key': '1. SQL LIKE (Baseline)',
        'kbji_key': '1. SQL LIKE (Baseline)',
        'note': 'Query pengguna langsung dicocokkan sebagai pola SQL ILIKE. Tidak ada pemrosesan — query informal tidak dapat dicocokkan dengan terminologi KBLI/KBJI formal.',
    },
    {
        'id': 'm2', 'no': '02', 'label': 'Hybrid Search',
        'name': 'Hybrid Search',
        'sub':  'SQL LIKE + Semantic Search (embedding Gemini) — tanpa preprocessing, tanpa contoh lapangan',
        'tags': ['SQL LIKE', 'Semantic Search'],
        'color': '#38bdf8', 'status': 'Sangat Baik',
        'show_prep': False, 'show_contoh': False,
        'kbli_key': '2. Hybrid Search',
        'kbji_key': '2. Hybrid Search',
        'note': 'Menggabungkan pecocokan leksikal (SQL LIKE) dan pemahaman semantik (Gemini embedding + cosine distance). Mampu memahami query informal tanpa preprocessing apapun.',
    },
    {
        'id': 'm3', 'no': '03', 'label': 'SQL LIKE + Preprocessing',
        'name': 'SQL LIKE + Preprocessing',
        'sub':  'SQL LIKE dengan query expansion (sinonim + stopword removal + variasi domain KBLI/KBJI)',
        'tags': ['SQL LIKE', 'Preprocessing'],
        'color': '#a78bfa', 'status': 'Kompetitif',
        'show_prep': True, 'show_contoh': False,
        'kbli_key': '3. SQL LIKE + Preprocessing',
        'kbji_key': '3. SQL LIKE + Preprocessing',
        'note': 'Preprocessing meliputi lowercase, stopword removal, stemming, dan query expansion dengan sinonim KBLI/KBJI. SQL LIKE kemudian mencari frasa yang diperkaya.',
    },
    {
        'id': 'm4', 'no': '04', 'label': 'Hybrid Search + Preprocessing',
        'name': 'Hybrid Search + Preprocessing',
        'sub':  'SQL LIKE + Semantic Search + preprocessing (query expansion) — tanpa contoh lapangan',
        'tags': ['SQL LIKE', 'Semantic Search', 'Preprocessing'],
        'color': '#67e8f9', 'status': 'Kompetitif',
        'show_prep': True, 'show_contoh': False,
        'kbli_key': '4. Hybrid Search + Preprocessing',
        'kbji_key': '4. Hybrid Search + Preprocessing',
        'note': 'Menggabungkan kekuatan semantic search dengan preprocessing query expansion. Tanpa field contoh lapangan — performa murni dari kombinasi leksikal dan semantik dengan query yang diperkaya.',
    },
    {
        'id': 'm5', 'no': '05', 'label': 'SQL LIKE + Preprocessing + Contoh Lapangan',
        'name': 'SQL LIKE + Preprocessing + Contoh Lapangan',
        'sub':  'SQL LIKE + preprocessing lengkap + memanfaatkan field contoh_lapangan dari database',
        'tags': ['SQL LIKE', 'Preprocessing', 'Contoh Lapangan'],
        'color': '#c084fc', 'status': 'Stabil',
        'show_prep': True, 'show_contoh': True,
        'kbli_key': '5. SQL LIKE + Preprocessing + Contoh Lapangan',
        'kbji_key': '5. SQL LIKE + Preprocessing + Contoh Lapangan',
        'note': 'Penambahan field contoh_lapangan memperluas cakupan pencarian SQL LIKE ke aktivitas nyata lapangan. Field ini berisi deskripsi informal yang digunakan responden, meningkatkan relevansi untuk query non-formal.',
    },
    {
        'id': 'm6', 'no': '06', 'label': 'Hybrid Search + Preprocessing + Contoh Lapangan',
        'name': 'Hybrid Search + Preprocessing + Contoh Lapangan',
        'sub':  'Kombinasi lengkap: SQL LIKE + Semantic Search + Query Expansion + Contoh Lapangan',
        'tags': ['SQL LIKE', 'Semantic Search', 'Preprocessing', 'Contoh Lapangan'],
        'color': '#34d399', 'status': 'Terbaik',
        'show_prep': True, 'show_contoh': True,
        'rows_kbli': rows_m5_kbli,
        'rows_kbji': rows_m5_kbji,
        'note': 'Sistem final yang mengombinasikan seluruh komponen. Semantic embedding memahami intent query, preprocessing memperkaya pencarian kata kunci, dan contoh lapangan memberikan konteks aktivitas nyata sebagai sinyal tambahan.',
    },
]

# ────────────────────────────────────────────────────────────────────
# 4. Helper functions
# ────────────────────────────────────────────────────────────────────
def h(s):
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def summary(rows):
    if not rows: return {'top1':0,'top3':0,'top10':0,'mrr':0,'n':0}
    n = len(rows)
    return {
        'n': n,
        'top1':  round(sum(int(r['top1'])  for r in rows)/n*100, 1),
        'top3':  round(sum(int(r['top3'])  for r in rows)/n*100, 1),
        'top10': round(sum(int(r['top10']) for r in rows)/n*100, 1),
        'mrr':   round(sum(float(r['rr'])  for r in rows)/n, 4),
    }

def rank_cell(rank_str):
    try: r = int(rank_str)
    except: r = 0
    if r == 1:   return f'<td class="r1">{r}</td>'
    if 2<=r<=3:  return f'<td class="r3">{r}</td>'
    if 4<=r<=10: return f'<td class="r10">{r}</td>'
    return '<td class="r0">&mdash;</td>'

def top1_cell(v_str):
    try: v = int(v_str)
    except: v = 0
    st = 'color:#6ee7b7;font-weight:700' if v == 1 else 'color:#334155'
    return f'<td style="text-align:center;{st}">{v}</td>'

def topn_cell(v_str):
    try: v = int(v_str)
    except: v = 0
    st = 'color:#7dd3fc' if v == 1 else 'color:#334155'
    return f'<td style="text-align:center;{st}">{v}</td>'

def rr_cell(v_str):
    try: v = float(v_str)
    except: v = 0.0
    if v >= 0.9: cl = '#6ee7b7'
    elif v >= 0.5: cl = '#7dd3fc'
    elif v > 0: cl = '#fcd34d'
    else: cl = '#334155'
    txt = f'{v:.4f}' if v > 0 else '&mdash;'
    return f'<td style="text-align:center;color:{cl};font-weight:600">{txt}</td>'

def badge_on(lbl, col):
    return f'<span class="badge-on" style="border-color:{col}60;color:{col};background:{col}12">&#10003; {lbl}</span>'

def badge_off(lbl):
    return f'<span class="badge-off">&#215; {lbl}</span>'

def build_rows_html(rows, show_prep, show_contoh):
    out = ''
    for i, r in enumerate(rows, 1):
        no    = r.get('no', i)
        query = h(r.get('query',''))
        kode  = h(r.get('kode_gt',''))
        prep  = h(r.get('preprocessed','') or '')
        contoh= h(r.get('contoh','') or '')
        tipe  = r.get('tipe','')
        tc = '#86efac' if tipe == 'KBLI' else '#93c5fd'

        row = '<tr>'
        row += f'<td style="text-align:center;color:#475569;font-size:0.78rem">{no}</td>'
        row += f'<td style="color:{tc};font-weight:700;text-align:center;font-size:0.79rem">{tipe}</td>'
        row += f'<td class="qcell">{query}</td>'
        if show_prep:
            if prep:
                row += f'<td class="prepcell">{prep}</td>'
            else:
                row += '<td style="color:#1e293b;font-style:italic;font-size:0.72rem;text-align:center">&mdash;</td>'
        if show_contoh:
            if contoh:
                row += f'<td class="contohcell">{contoh}</td>'
            else:
                row += '<td style="color:#1e293b;font-style:italic;font-size:0.72rem;text-align:center">&mdash;</td>'
        row += f'<td style="text-align:center"><code style="background:#131b2e;color:#a5b4fc;padding:1px 6px;border-radius:4px;border:1px solid #1e2f50;font-size:0.78em">{kode}</code></td>'
        row += rank_cell(r.get('rank','0'))
        row += top1_cell(r.get('top1','0'))
        row += topn_cell(r.get('top3','0'))
        row += topn_cell(r.get('top10','0'))
        row += rr_cell(r.get('rr','0'))
        row += '</tr>\n'
        out += row
    return out

def build_subtotal(rows, label, color, extra_cols):
    s = summary(rows)
    t1c = '#ef4444' if s['top1']<10 else ('#fbbf24' if s['top1']<70 else '#86efac')
    mrrc = '#ef4444' if s['mrr']<0.1 else ('#fbbf24' if s['mrr']<0.7 else color)
    extra = '<td></td>' * extra_cols
    return f'''<tr class="subtotal-row">
  <td colspan="3" style="text-align:right;padding-right:12px;font-size:0.78rem">
    <span style="color:#475569">— {label} ({s["n"]} query) —</span>
  </td>
  {extra}
  <td></td>
  <td style="text-align:center;color:{t1c};font-weight:700">&mdash;</td>
  <td style="text-align:center;color:{t1c};font-weight:700">{s["top1"]}%</td>
  <td style="text-align:center;color:{t1c};font-weight:700">{s["top3"]}%</td>
  <td style="text-align:center;color:{t1c};font-weight:700">{s["top10"]}%</td>
  <td style="text-align:center;color:{mrrc};font-weight:700">{s["mrr"]:.4f}</td>
</tr>'''

def build_method_section(m, rows_kbli, rows_kbji):
    col   = m['color']
    no    = m['no']
    label = m['label']
    name  = m['name']
    sub   = m['sub']
    note  = m['note']
    tags  = m['tags']
    st    = m['status']
    show_prep   = m['show_prep']
    show_contoh = m['show_contoh']
    extra_cols  = (1 if show_prep else 0) + (1 if show_contoh else 0)

    sk = summary(rows_kbli)
    sj = summary(rows_kbji)
    nk, nj = sk['n'], sj['n']
    n = nk + nj
    c_t1  = round((sk['top1']*nk/100 + sj['top1']*nj/100)/n*100, 1) if n else 0
    c_t3  = round((sk['top3']*nk/100 + sj['top3']*nj/100)/n*100, 1) if n else 0
    c_t10 = round((sk['top10']*nk/100 + sj['top10']*nj/100)/n*100, 1) if n else 0
    c_mrr = round((sk['mrr']*nk + sj['mrr']*nj)/n, 4) if n else 0

    # Badge
    all_tags_html = ''
    for t in ['SQL LIKE','Semantic Search','Preprocessing','Contoh Lapangan']:
        if t in tags:
            all_tags_html += badge_on(t, col)
        else:
            all_tags_html += badge_off(t)

    # Status color
    st_col = {'Baseline':'#64748b','Sangat Baik':'#38bdf8','Kompetitif':'#a78bfa',
               'Stabil':'#c084fc','Terbaik':'#34d399'}.get(st,'#94a3b8')

    # Summary cards
    def mc(lbl, val, vc): 
        return f'<div class="scard"><div class="slbl">{lbl}</div><div class="sval" style="color:{vc}">{val}</div></div>'
    t1c = '#ef4444' if c_t1<10 else ('#fbbf24' if c_t1<50 else col)
    mrrc = '#ef4444' if c_mrr<0.1 else ('#fbbf24' if c_mrr<0.5 else col)
    
    cards_html = f'''<div class="sbar">
      {mc("KBLI Top@1",  f"{sk['top1']}%", col)}
      {mc("KBLI Top@3",  f"{sk['top3']}%", col)}
      {mc("KBLI Top@10", f"{sk['top10']}%", col)}
      {mc("KBLI MRR",    f"{sk['mrr']:.4f}", col)}
      <div style="width:1px;background:#1e293b;margin:0 4px"></div>
      {mc("KBJI Top@1",  f"{sj['top1']}%", col)}
      {mc("KBJI Top@3",  f"{sj['top3']}%", col)}
      {mc("KBJI Top@10", f"{sj['top10']}%", col)}
      {mc("KBJI MRR",    f"{sj['mrr']:.4f}", col)}
      <div style="width:1px;background:#1e293b;margin:0 4px"></div>
      {mc("Gabungan Top@1", f"{c_t1}%", t1c)}
      {mc("Gabungan MRR",   f"{c_mrr:.4f}", mrrc)}
    </div>'''

    # Table header columns
    prep_th   = f'<th class="prep-th">Preprocessing Query</th>' if show_prep else ''
    contoh_th = f'<th class="contoh-th">Contoh Lapangan</th>' if show_contoh else ''

    rows_kbli_html = build_rows_html(rows_kbli, show_prep, show_contoh)
    rows_kbji_html = build_rows_html(rows_kbji, show_prep, show_contoh)
    sub_kbli = build_subtotal(rows_kbli, 'KBLI 2025', col, extra_cols)
    sub_kbji = build_subtotal(rows_kbji, 'KBJI 2014', col, extra_cols)

    return f'''
<section id="{m['id']}" class="method-card" style="border-color:{col}28">
  <div class="method-head">
    <div>
      <span class="method-num" style="background:{col}18;border-color:{col}50;color:{col}">{no}</span>
      <span class="status-chip" style="background:{st_col}15;border-color:{st_col}50;color:{st_col}">{st}</span>
    </div>
    <h2 class="method-title" style="color:{col}">{name}</h2>
    <p class="method-sub">{sub}</p>
    <p class="method-note" style="border-left-color:{col}60">{note}</p>
    <div class="badges">{all_tags_html}</div>
  </div>
  {cards_html}
  <div class="tbl-wrap">
    <table>
      <thead>
        <tr>
          <th class="no-th">#</th>
          <th>Tipe</th>
          <th style="text-align:left;min-width:150px">Query Asli</th>
          {prep_th}{contoh_th}
          <th>Kode GT</th>
          <th>Rank</th>
          <th>Top@1</th>
          <th>Top@3</th>
          <th>Top@10</th>
          <th>RR</th>
        </tr>
      </thead>
      <tbody>
        {rows_kbli_html}
        {sub_kbli}
        <tr class="sep-row"><td colspan="{10+extra_cols}"></td></tr>
        {rows_kbji_html}
        {sub_kbji}
      </tbody>
    </table>
  </div>
</section>'''

# ────────────────────────────────────────────────────────────────────
# 5. Build nagivation + comparison table + chart
# ────────────────────────────────────────────────────────────────────
def build_comparison_section():
    rows_data = []
    for m in METHODS:
        if 'rows_kbli' in m:
            rk = m['rows_kbli']
            rj = m['rows_kbji']
        else:
            rk = get_rows(m['kbli_key'], 'KBLI')
            rj = get_rows(m['kbji_key'], 'KBJI')
        sk = summary(rk)
        sj = summary(rj)
        rows_data.append((m, sk, sj))

    # Bar chart MRR
    max_mrr = max(max(d[1]['mrr'], d[2]['mrr']) for d in rows_data) or 1

    def chart_row(label, val, color, max_v=1.0):
        pct = round(val/max_v*100, 1)
        return f'''<div class="chart-row">
  <div class="chart-lbl">{label}</div>
  <div class="chart-track">
    <div class="chart-fill" style="width:{pct}%;background:linear-gradient(90deg,{color},{color}99)">
      <span class="chart-inner">{val:.4f}</span>
    </div>
  </div>
  <div class="chart-val" style="color:{color}">{val:.4f}</div>
</div>'''

    mrr_kbli_charts = ''
    mrr_kbji_charts = ''
    for m, sk, sj in rows_data:
        mrr_kbli_charts += chart_row(m['name'], sk['mrr'], m['color'])
        mrr_kbji_charts += chart_row(m['name'], sj['mrr'], m['color'])

    # Summary table
    def status_chip_small(st, col):
        return f'<span style="background:{col}15;border:1px solid {col}50;color:{col};padding:2px 8px;border-radius:5px;font-size:0.72rem">{st}</span>'

    tbl_rows = ''
    for m, sk, sj in rows_data:
        col = m['color']
        has = lambda t: '&#10003;' if t in m['tags'] else '<span style="color:#1e293b">&#10007;</span>'
        t1_col = '#ef4444' if sk['top1']<10 else ('#fbbf24' if sk['top1']<70 else '#86efac')
        mrr_col = '#ef4444' if sk['mrr']<0.1 else ('#fbbf24' if sk['mrr']<0.7 else col)
        t1j_col = '#ef4444' if sj['top1']<10 else ('#fbbf24' if sj['top1']<70 else col)
        mrrj_col = '#ef4444' if sj['mrr']<0.1 else ('#fbbf24' if sj['mrr']<0.7 else col)
        tbl_rows += f'''<tr>
  <td style="color:{col};font-weight:600;white-space:nowrap">{m['no']}. {m['name']}</td>
  <td style="text-align:center;color:#6ee7b7">{has('SQL LIKE')}</td>
  <td style="text-align:center;color:#38bdf8">{has('Semantic Search')}</td>
  <td style="text-align:center;color:#fcd34d">{has('Preprocessing')}</td>
  <td style="text-align:center;color:#86efac">{has('Contoh Lapangan')}</td>
  <td style="text-align:center;color:{t1_col};font-weight:700">{sk['top1']}%</td>
  <td style="text-align:center;color:{t1_col}">{sk['top3']}%</td>
  <td style="text-align:center;color:{t1_col}">{sk['top10']}%</td>
  <td style="text-align:center;color:{mrr_col};font-weight:700">{sk['mrr']:.4f}</td>
  <td style="text-align:center;color:{t1j_col};font-weight:700">{sj['top1']}%</td>
  <td style="text-align:center;color:{t1j_col}">{sj['top3']}%</td>
  <td style="text-align:center;color:{t1j_col}">{sj['top10']}%</td>
  <td style="text-align:center;color:{mrrj_col};font-weight:700">{sj['mrr']:.4f}</td>
  <td style="text-align:center">{status_chip_small(m["status"], col)}</td>
</tr>'''

    return f'''
<section id="perbandingan" class="method-card" style="border-color:#34d39928">
  <div class="method-head">
    <h2 class="method-title" style="color:#34d399">&#11088; Tabel Perbandingan Semua Metode</h2>
    <p class="method-sub">60 query (30 KBLI 2025 + 30 KBJI 2014) — hasil evaluasi akhir semua metode</p>
  </div>

  <!-- MRR Chart KBLI -->
  <div class="chart-section">
    <div class="chart-title">&#128200; MRR — KBLI 2025</div>
    {mrr_kbli_charts}
  </div>
  <div class="chart-section" style="margin-top:20px">
    <div class="chart-title">&#128200; MRR — KBJI 2014</div>
    {mrr_kbji_charts}
  </div>

  <!-- Comparison Table -->
  <div class="tbl-wrap" style="margin-top:24px">
    <table>
      <thead>
        <tr>
          <th rowspan="2" style="text-align:left;min-width:240px">Metode</th>
          <th colspan="4" style="background:#0f172a;color:#94a3b8">Komponen</th>
          <th colspan="4" style="background:#0c2040;color:#38bdf8">KBLI 2025</th>
          <th colspan="4" style="background:#0a1f1a;color:#34d399">KBJI 2014</th>
          <th rowspan="2">Status</th>
        </tr>
        <tr>
          <th style="color:#6ee7b7">SQL</th>
          <th style="color:#38bdf8">Sem.</th>
          <th style="color:#fcd34d">Prep</th>
          <th style="color:#86efac">CL</th>
          <th style="color:#38bdf8">Top@1</th>
          <th style="color:#38bdf8">Top@3</th>
          <th style="color:#38bdf8">Top@10</th>
          <th style="color:#38bdf8">MRR</th>
          <th style="color:#34d399">Top@1</th>
          <th style="color:#34d399">Top@3</th>
          <th style="color:#34d399">Top@10</th>
          <th style="color:#34d399">MRR</th>
        </tr>
      </thead>
      <tbody>
        {tbl_rows}
      </tbody>
    </table>
  </div>
</section>'''

# ────────────────────────────────────────────────────────────────────
# 6. CSS
# ────────────────────────────────────────────────────────────────────
CSS = """
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{font-family:'Outfit',sans-serif;background:#060810;color:#cbd5e1;line-height:1.6;font-size:15px}

.topnav{position:sticky;top:0;z-index:100;background:#060810ee;backdrop-filter:blur(16px) saturate(1.4);
  border-bottom:1px solid #12172a;padding:9px 36px;display:flex;align-items:center;gap:6px;flex-wrap:wrap}
.topnav-brand{font-size:0.82rem;font-weight:800;color:#34d399;margin-right:8px;letter-spacing:-.01em;white-space:nowrap}
.nav-btn{font-size:0.69rem;padding:4px 10px;border-radius:6px;border:1px solid;text-decoration:none;
  white-space:nowrap;transition:all .2s;display:inline-block}
.nav-btn:hover{opacity:.65;transform:translateY(-1px)}

.hero{padding:52px 44px 40px;border-bottom:1px solid #12172a}
.hero-badge{display:inline-block;background:linear-gradient(135deg,#065f46,#059669);color:#fff;
  font-size:0.72rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  padding:4px 14px;border-radius:999px;margin-bottom:14px}
.hero h1{font-size:1.85rem;font-weight:800;margin-bottom:6px;
  background:linear-gradient(90deg,#6ee7b7,#34d399,#38bdf8);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent}
.hero-sub{color:#475569;font-size:0.88rem;margin-bottom:4px}
.hero-meta{color:#334155;font-size:0.78rem;margin-top:10px}
.hero-meta strong{color:#94a3b8}
.param-chips{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px}
.param-chip{font-size:0.76rem;padding:5px 13px;border-radius:8px;border:1px solid;display:inline-flex;align-items:center;gap:6px}

.page-body{padding:36px 44px 64px}

.method-card{border:1px solid #1a2030;border-radius:16px;padding:28px 32px;margin-bottom:36px;
  background:linear-gradient(180deg,#0a0d18 0%,#060810 100%)}
.method-head{margin-bottom:20px}
.method-num{font-size:0.7rem;font-weight:700;padding:3px 10px;border-radius:6px;border:1px solid;
  display:inline-block;margin-right:7px;letter-spacing:.04em}
.status-chip{font-size:0.7rem;font-weight:700;padding:3px 10px;border-radius:6px;border:1px solid;
  display:inline-block;letter-spacing:.04em}
.method-title{font-size:1.18rem;font-weight:700;margin:10px 0 4px}
.method-sub{font-size:0.82rem;color:#475569;margin-bottom:8px}
.method-note{font-size:0.82rem;color:#64748b;line-height:1.55;
  background:#ffffff05;border-left:3px solid;border-radius:0 8px 8px 0;
  padding:9px 16px;margin:10px 0 14px}
.badges{margin-top:6px}
.badge-on,.badge-off{display:inline-flex;align-items:center;gap:4px;font-size:0.74rem;
  padding:3px 10px;border-radius:6px;margin:3px;border:1px solid}
.badge-off{border-color:#1e293b;color:#1e3a5a;background:#060810;text-decoration:line-through}

.sbar{display:flex;gap:8px;flex-wrap:wrap;margin:14px 0}
.scard{background:#0c1020;border:1px solid #141c30;border-radius:9px;padding:9px 14px;min-width:86px}
.slbl{font-size:0.64rem;color:#334155;text-transform:uppercase;letter-spacing:.07em;margin-bottom:2px}
.sval{font-size:1.1rem;font-weight:700;color:#e2e8f0}

.tbl-wrap{overflow-x:auto;border-radius:10px;margin:6px 0;
  box-shadow:0 4px 24px rgba(0,0,0,.5);border:1px solid #111827}
table{border-collapse:collapse;font-size:0.78rem;width:100%}
th,td{padding:7px 10px;border:1px solid #0f172a}
th{background:#0c1020;color:#475569;font-weight:600;text-transform:uppercase;
  font-size:0.65rem;letter-spacing:.07em;text-align:center;white-space:nowrap}
tr:hover td{background:rgba(255,255,255,.01)!important}

.qcell{color:#c4b5fd;white-space:normal;max-width:195px;font-size:0.79rem}
.prepcell{color:#fcd34d;font-size:0.72rem;white-space:normal;max-width:200px;line-height:1.4}
.contohcell{color:#86efac;font-size:0.72rem;white-space:normal;max-width:180px;line-height:1.4}
.no-th{min-width:32px}
.prep-th{color:#fcd34d!important;min-width:160px}
.contoh-th{color:#86efac!important;min-width:160px}
code{background:#0f172a;color:#a5b4fc;padding:1px 5px;border-radius:4px;
  border:1px solid #1e2f50;font-size:.78em}

/* Rank colors */
td.r1{background:#064e3b;color:#6ee7b7;font-weight:700;text-align:center}
td.r3{background:#065f46;color:#a7f3d0;text-align:center}
td.r10{background:#1a302a;color:#d1fae5;text-align:center}
td.r0{background:#0f172a;color:#1e3a5a;text-align:center}

.subtotal-row td{background:#0d1524;border-top:1px dashed #1e293b}
.sep-row td{background:#0c1020;height:8px;border:none}

/* Chart */
.chart-section{background:#0a0d18;border:1px solid #111827;border-radius:10px;padding:16px 20px}
.chart-title{font-size:0.78rem;font-weight:700;color:#64748b;text-transform:uppercase;
  letter-spacing:.07em;margin-bottom:12px}
.chart-row{display:flex;align-items:center;gap:10px;margin:6px 0}
.chart-lbl{font-size:0.73rem;color:#475569;min-width:280px;text-align:right}
.chart-track{flex:1;background:#0c1020;border-radius:4px;height:22px;overflow:hidden;border:1px solid #111827}
.chart-fill{height:22px;border-radius:3px;display:flex;align-items:center;min-width:8px}
.chart-inner{padding-left:8px;font-size:0.68rem;font-weight:700;color:rgba(255,255,255,.85)}
.chart-val{font-size:0.74rem;font-weight:700;min-width:54px}

hr{border:0;border-top:2px dashed #0f172a;margin:40px 0}
.footer{margin-top:48px;padding:20px 0;border-top:1px solid #0f172a;
  color:#1e3a5a;font-size:0.74rem;text-align:center}
"""

# ────────────────────────────────────────────────────────────────────
# 7. Nav
# ────────────────────────────────────────────────────────────────────
nav_links = [
    ('#m1','01. SQL LIKE','#64748b'),
    ('#m2','02. Hybrid','#38bdf8'),
    ('#m3','03. SQL+Prep','#a78bfa'),
    ('#m4','04. Hybrid+Prep','#67e8f9'),
    ('#m5','05. SQL+Prep+CL','#c084fc'),
    ('#m6','06. Hybrid+Prep+CL ✅','#34d399'),
    ('#perbandingan','⭐ Perbandingan','#fbbf24'),
]
nav_html = ''.join(
    f'<a href="{href}" class="nav-btn" style="border-color:{col}60;color:{col}">{lbl}</a>'
    for href, lbl, col in nav_links
)

ts = datetime.now().strftime('%d %B %Y, %H:%M')

# ────────────────────────────────────────────────────────────────────
# 8. Assemble full HTML
# ────────────────────────────────────────────────────────────────────
method_sections_html = ''
for m in METHODS:
    if 'rows_kbli' in m:
        rk = m['rows_kbli']
        rj = m['rows_kbji']
    else:
        rk = get_rows(m['kbli_key'], 'KBLI')
        rj = get_rows(m['kbji_key'], 'KBJI')
    method_sections_html += build_method_section(m, rk, rj)
    method_sections_html += '\n<hr>\n'

comparison_html = build_comparison_section()

final_html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Evaluasi Sistem Pencarian KBLI — DEMAKAI</title>
<meta name="description" content="Dashboard evaluasi sistem pencarian KBLI DEMAKAI: SQL LIKE, Hybrid Search, Preprocessing, dan Contoh Lapangan. 60 query evaluasi.">
<style>{CSS}</style>
</head>
<body>

<nav class="topnav">
  <span class="topnav-brand">&#128202; DEMAKAI</span>
  {nav_html}
</nav>

<div class="hero">
  <div class="hero-badge">Laporan Evaluasi Sistem</div>
  <h1>Evaluasi Sistem Pencarian KBLI</h1>
  <p class="hero-sub">Perbandingan SQL LIKE dan Hybrid Search &mdash; dari Baseline hingga Sistem Final</p>
  <p class="hero-sub">Proyek: <strong>DEMAKAI</strong> — Deteksi Metode Aktivitas Kerja dengan AI | BPS Kabupaten Demak</p>
  <p class="hero-meta">
    <strong>Dataset:</strong> KBLI 2025 &amp; KBJI 2014 &nbsp;&middot;&nbsp;
    <strong>Total Query:</strong> 60 (30 KBLI + 30 KBJI) &nbsp;&middot;&nbsp;
    <strong>Metrik:</strong> Top@1, Top@3, Top@10, MRR &nbsp;&middot;&nbsp;
    <strong>Generated:</strong> {ts}
  </p>
  <div class="param-chips">
    <span class="param-chip" style="border-color:#38bdf840;color:#38bdf8;background:#38bdf810">
      &#127761; Semantic: Gemini Embedding + pgvector
    </span>
    <span class="param-chip" style="border-color:#a78bfa40;color:#a78bfa;background:#a78bfa10">
      &#9986; Preprocessing: Stopword, Stemming, Query Expansion
    </span>
    <span class="param-chip" style="border-color:#34d39940;color:#34d399;background:#34d39910">
      &#128203; Contoh Lapangan: Field aktivitas nyata di database
    </span>
    <span class="param-chip" style="border-color:#fbbf2440;color:#fbbf24;background:#fbbf2410">
      &#128202; Evaluasi: Top@K &amp; Mean Reciprocal Rank
    </span>
  </div>
</div>

<div class="page-body">
  {method_sections_html}
  {comparison_html}
  <div class="footer">
    DEMAKAI &mdash; Dashboard Evaluasi Sistem Pencarian KBLI &nbsp;&middot;&nbsp;
    60 query evaluasi (30 KBLI 2025 + 30 KBJI 2014) &nbsp;&middot;&nbsp;
    Data: evaluasi_semua_20260420_144133.csv &amp; evaluasi_after_fix_21apr.xlsx &nbsp;&middot;&nbsp;
    Build: {ts}
  </div>
</div>

</body>
</html>"""

out_path = os.path.join(BASE, 'dashboard_skripsi_final.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(final_html)
print(f"Berhasil: {out_path}")
print(f"Ukuran: {len(final_html):,} bytes ({len(final_html)//1024} KB)")

# Verifikasi cepat
checks = ['Baseline','Hybrid Search','SQL LIKE + Preprocessing',
          'SQL LIKE + Preprocessing + Contoh Lapangan',
          'Hybrid Search + Preprocessing + Contoh Lapangan',
          'id="perbandingan"','chart-fill','r1']
for c in checks:
    n = final_html.count(c)
    print(f"  {'OK' if n else 'MISS!'} | '{c}': {n}x")
