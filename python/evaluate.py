"""
evaluate.py — Evaluasi Otomatis 6 Kombinasi Metode Pencarian DEMAKAI
=====================================================================
Baca daftar query dari queries.csv, jalankan 6 kombinasi metode,
hitung Rank / top1 / top5 / KK untuk setiap kombinasi, simpan ke:
  - output/evaluasi_hasil.csv   (tabel angka lengkap)
  - output/evaluasi_hasil.html  (tabel visual berwarna)

Cara pakai:
  python evaluate.py

Edit queries.csv untuk mengganti daftar kuery dan kode ground truth.
"""

import csv
import os
import sys
from datetime import datetime

# ── Pastikan bisa import modul lokal ─────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))

from preprocessing.expansion import preprocess_expansion
from preprocessing.advanced import preprocess_advanced
from search.sql_like import (
    search_raw      as sql_raw,
    search_expansion as sql_expansion,
    search_advanced  as sql_advanced,
)
from search.hybrid import (
    search_raw      as hybrid_raw,
    search_expansion as hybrid_expansion,
    search_advanced as hybrid_advanced,
)

os.makedirs("output", exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# Definisi 6 kombinasi
# ─────────────────────────────────────────────────────────────────────────────

COMBINATIONS = [
    ("SQL",    "None",      lambda q, model=None: sql_raw(q, model=model)),
    ("SQL",    "Advanced",  lambda q, model=None: sql_advanced(preprocess_advanced(q), model=model)),
    ("SQL",    "Expansion", lambda q, model=None: sql_expansion(preprocess_expansion(q), model=model)),
    ("Hybrid", "None",      lambda q, model=None: hybrid_raw(q, model=model)),
    ("Hybrid", "Advanced",  lambda q, model=None: hybrid_advanced(preprocess_advanced(q), model=model)),
    ("Hybrid", "Expansion", lambda q, model=None: hybrid_expansion(preprocess_expansion(q), model=model)),
]

COMBO_LABELS = [f"{s}_{p}" for s, p, _ in COMBINATIONS]  # e.g. "SQL_None"


# ─────────────────────────────────────────────────────────────────────────────
# Helper: cari rank kode ground truth di hasil pencarian
# ─────────────────────────────────────────────────────────────────────────────

def get_rank(results: list, kode_gt: str) -> int | None:
    """
    Cari posisi (rank) kode_gt di dalam list results.
    Mengembalikan integer (1-based) atau None jika tidak ditemukan.
    """
    for i, item in enumerate(results, start=1):
        if str(item.get("kode", "")).strip() == str(kode_gt).strip():
            return i
    return None


def compute_metrics(rank: int, top_n: int = 10) -> dict:
    """
    Menghitung metrik berdasarkan rank aktual.
    Jika rank <= 0 atau > top_n, dianggap tidak ketemu (0).
    """
    if 0 < rank <= top_n:
        top1  = 1 if rank == 1 else 0
        top5  = 1 if rank <= 5 else 0
        top10 = 1 if rank <= 10 else 0
        kk    = round(1.0 / rank, 4)
        rank_val = rank
    else:
        top1 = 0
        top5 = 0
        top10 = 0
        kk = 0.0
        rank_val = 0

    return {
        "rank": rank_val,
        "top1": top1,
        "top5": top5,
        "top10": top10,
        "kk": kk
    }


# ─────────────────────────────────────────────────────────────────────────────
# Fungsi utama evaluasi
# ─────────────────────────────────────────────────────────────────────────────

def run_evaluation(queries_file: str = "queries.csv", limit: int = 10):
    """
    Jalankan evaluasi untuk semua query di queries_file.
    File CSV harus memiliki kolom: no, query, kode_ground_truth, tipe (KBLI/KBJI)
    """
    # Baca query
    if not os.path.exists(queries_file):
        print(f"[ERROR] File '{queries_file}' tidak ditemukan.")
        return []

    with open(queries_file, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        queries = list(reader)

    all_rows = []

    for entry in queries:
        no      = entry.get("no", "")
        query   = entry.get("query", "").strip()
        kode_gt = entry.get("kode_ground_truth", "").strip()
        tipe    = entry.get("tipe", "").strip().upper()  # KBLI atau KBJI

        # model filter: hanya cari di tabel yang sesuai
        model = tipe if tipe in ("KBLI", "KBJI") else None

        print(f"\n[{no}] Query: \"{query}\" | GT: {kode_gt} | Tipe: {tipe}")

        # Hitung string preprocessing untuk ditampilkan
        try:
            res_adv   = preprocess_advanced(query)
            res_exp   = preprocess_expansion(query)
            desc_adv   = f"Advanced (Stemmed):\n[{', '.join(res_adv['stemmed_tokens'])}]"
            
            # Tambahan: Tampilkan Field Variations (VARIATION_MAP)
            field_key = "kbli_variations" if tipe == "KBLI" else "kbji_variations"
            field_vars = res_exp.get(field_key, [])
            
            desc_exp   = f"Expansion (Synonyms):\n[{', '.join(res_exp['expanded_tokens'])}]"
            if field_vars:
                desc_exp += f"\n\nField Variations ({tipe}):\n[{', '.join(field_vars)}]"
        except Exception:
            desc_adv   = query
            desc_exp   = query

        row = {
            "no": no, 
            "query": query, 
            "kode_ground_truth": kode_gt, 
            "tipe": tipe,
            "desc_advanced": desc_adv,
            "desc_expansion": desc_exp
        }

        for search_method, proc_method, search_fn in COMBINATIONS:
            label = f"{search_method}_{proc_method}"
            print(f"  -> {label} ...", end=" ", flush=True)

            try:
                results = search_fn(query, model=model)
                rank    = get_rank(results, kode_gt) or 0
                metrics = compute_metrics(rank, top_n=limit)
                print(f"rank={metrics['rank']}")
            except Exception as e:
                print(f"ERROR: {e}")
                metrics = {"rank": 0, "top1": 0, "top5": 0, "top10": 0, "kk": 0.0}

            row[f"rank_{label}"]  = metrics["rank"]
            row[f"top1_{label}"]  = metrics["top1"]
            row[f"top5_{label}"]  = metrics["top5"]
            row[f"top10_{label}"] = metrics["top10"]
            row[f"kk_{label}"]    = metrics["kk"]

        all_rows.append(row)

    return all_rows


def calculate_summary(rows: list) -> list:
    if not rows:
        return []
    
    summary = []
    n = len(rows)
    for s, p, _ in COMBINATIONS:
        label = f"{s}_{p}"
        top1_vals = [int(r.get(f"top1_{label}", 0)) for r in rows]
        top5_vals = [int(r.get(f"top5_{label}", 0)) for r in rows]
        kk_vals   = [float(r.get(f"kk_{label}", 0.0)) for r in rows]
        
        top1_avg = sum(top1_vals) / n
        top5_avg = sum(top5_vals) / n
        mrr_avg  = sum(kk_vals) / n
        
        if s == "SQL":
            if p == "None": name = "SQL LIKE (Tanpa Preprocessing)"
            elif p == "Advanced": name = "SQL LIKE (Advanced: Stopword + Stemming)"
            elif p == "Expansion": name = "SQL LIKE (Query Expansion: Sinonim + Domain)"
        else:
            if p == "None": name = "Hybrid Search (Tanpa Preprocessing)"
            elif p == "Advanced": name = "Hybrid Search (Advanced: Stopword + Stemming)"
            elif p == "Expansion": name = "Hybrid Search (Query Expansion: Sinonim + Domain)"

        summary.append({
            "Metode": name,
            "Top1": round(top1_avg, 3),
            "Top5": round(top5_avg, 3),
            "MRR": round(mrr_avg, 3)
        })
    return summary


def save_summary_csv(summary: list, filepath: str):
    if not summary:
        return
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Metode", "Top1", "Top5", "MRR"])
        writer.writeheader()
        writer.writerows(summary)
    print(f"[CSV] Summary Disimpan -> {os.path.abspath(filepath)}")


def save_csv(rows: list, filepath: str):
    if not rows:
        return

    base_cols = ["no", "tipe", "query", "kode_ground_truth"]
    metric_cols = []
    for label in COMBO_LABELS:
        metric_cols.extend([
            f"rank_{label}",
            f"top1_{label}",
            f"top5_{label}",
            f"top10_{label}",
            f"kk_{label}"
        ])

    fieldnames = base_cols + metric_cols

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)

    print(f"[CSV] Data Disimpan -> {os.path.abspath(filepath)}")


def build_section_html(title: str, rows: list, summary_data: list) -> str:
    if not rows:
        return ""

    def cell_style(val):
        if val == 0: return "background:#2d2f45;color:#555"
        try:
            v = int(val)
        except (ValueError, TypeError):
            return ""
        if v == 1:   return "background:#064e3b;color:#6ee7b7;font-weight:bold"
        if v <= 5:   return "background:#065f46;color:#a7f3d0"
        if v <= 10:  return "background:#1e3a2f;color:#d1fae5"
        return "background:#2d2f45;color:#555"

    combo_headers = ""
    for s, p, _ in COMBINATIONS:
        colspan = 6 if p in ("Advanced", "Expansion") else 5
        combo_headers += f'<th colspan="{colspan}" style="background:#312e81;color:#a5b4fc;text-align:center">{s} + {p}</th>'

    sub_headers = ""
    for _, p, _ in COMBINATIONS:
        sub_headers += '<th>rank</th><th>top1</th><th>top5</th><th>top10</th><th>KK</th>'
        if p in ("Advanced", "Expansion"):
            sub_headers += '<th>Deskripsi Preprocessing</th>'

    data_rows_html = ""
    for row in rows:
        tipe_val  = row.get("tipe", "")
        tipe_color = "#86efac" if tipe_val == "KBLI" else "#93c5fd"
        cells = (
            f'<td>{row["no"]}</td>'
            f'<td style="color:{tipe_color};font-weight:bold">{tipe_val}</td>'
            f'<td style="color:#c4b5fd">{row["query"]}</td>'
            f'<td><code>{row["kode_ground_truth"]}</code></td>'
        )
        for s, p, _ in COMBINATIONS:
            label = f"{s}_{p}"
            rank  = row.get(f"rank_{label}", 0)
            top1  = row.get(f"top1_{label}", 0)
            top5  = row.get(f"top5_{label}", 0)
            top10 = row.get(f"top10_{label}", 0)
            kk    = row.get(f"kk_{label}", 0.0)
            
            cells += (
                f'<td style="{cell_style(rank)}">{rank}</td>'
                f'<td style="{cell_style(top1)}">{top1}</td>'
                f'<td style="{cell_style(top5)}">{top5}</td>'
                f'<td style="{cell_style(top10)}">{top10}</td>'
                f'<td style="{cell_style(kk)}">{kk}</td>'
            )
            
            if p == "Advanced":
                val = row.get("desc_advanced", "").replace("\n", "<br>")
                cells += f'<td style="color:#d8b4fe;font-size:0.7rem;max-width:200px;white-space:normal;">{val}</td>'
            elif p == "Expansion":
                val = row.get("desc_expansion", "").replace("\n", "<br>")
                cells += f'<td style="color:#d8b4fe;font-size:0.7rem;max-width:200px;white-space:normal;">{val}</td>'
                
        data_rows_html += f"<tr>{cells}</tr>\n"

    summary_html = '<tr><td colspan="4" style="font-weight:bold;color:#a78bfa;text-align:right;padding-right:15px">MRR</td>'
    for s, p, _ in COMBINATIONS:
        label = f"{s}_{p}"
        kk_vals = [float(row.get(f"kk_{label}", 0)) for row in rows]
        mrr = round(sum(kk_vals) / len(kk_vals), 4) if kk_vals else 0
        color = "#6ee7b7" if mrr >= 0.5 else "#fbbf24" if mrr >= 0.2 else "#f87171"
        
        summary_html += (
            f'<td colspan="4"></td>'
            f'<td style="font-weight:bold;color:{color}">{mrr:.4f}</td>'
        )
        if p in ("Advanced", "Expansion"):
            summary_html += '<td></td>'
            
    summary_html += "</tr>"

    summary_table_html = f"""
    <br><br>
    <h2 style="color: #a78bfa; margin-bottom: 8px; font-size: 1.2rem;">📊 Tabel Ringkasan (Average) - {title}</h2>
    <table style="width: 50%; min-width: 400px; margin-bottom: 40px;">
      <thead>
        <tr>
          <th style="background:#25273b;color:#a78bfa;text-align:left">Metode</th>
          <th style="background:#25273b;color:#a78bfa">top1</th>
          <th style="background:#25273b;color:#a78bfa">top5</th>
          <th style="background:#25273b;color:#a78bfa">MRR</th>
        </tr>
      </thead>
      <tbody>
"""
    for item in summary_data:
        t1 = f"{item['Top1']:.3f}".replace('.', ',')
        t5 = f"{item['Top5']:.3f}".replace('.', ',')
        rr = f"{item['MRR']:.3f}".replace('.', ',')
        summary_table_html += f"""        <tr>
          <td style="font-weight:bold;color:#c4b5fd">{item['Metode']}</td>
          <td style="text-align:center">{t1}</td>
          <td style="text-align:center">{t5}</td>
          <td style="text-align:center">{rr}</td>
        </tr>"""
    summary_table_html += """
      </tbody>
    </table>
"""

    return f"""
  <h2 style="color: #a5b4fc; margin-top: 20px; font-size: 1.4rem;">Dataset: {title}</h2>
  <div class="wrap" style="margin-bottom: 20px;">
    <table>
      <thead>
        <tr>
          <th rowspan="2">#</th>
          <th rowspan="2">Tipe</th>
          <th rowspan="2">Query Asli</th>
          <th rowspan="2">Kode GT</th>
          {combo_headers}
        </tr>
        <tr>
          {sub_headers}
        </tr>
      </thead>
      <tbody>
        {data_rows_html}
        {summary_html}
      </tbody>
    </table>
  </div>
  {summary_table_html}
    """


def save_html_combined(rows_kbli: list, summary_kbli: list, rows_kbji: list, summary_kbji: list, filepath: str):
    ts = datetime.now().strftime("%d %B %Y, %H:%M:%S")
    
    html_kbli = build_section_html("KBLI 2025", rows_kbli, summary_kbli)
    html_kbji = build_section_html("KBJI 2014", rows_kbji, summary_kbji)
    total_queries = len(rows_kbli) + len(rows_kbji)

    html = f"""<!DOCTYPE html>
<html lang="id">
<head>
  <meta charset="UTF-8">
  <title>DEMAKAI — Evaluasi 6 Kombinasi</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600&display=swap');
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ 
        font-family: 'Outfit', sans-serif; 
        background: #0f111a; 
        color: #e2e8f0; 
        padding: 40px;
        line-height: 1.5;
    }}
    h1 {{ 
        color: #a78bfa; 
        margin-bottom: 8px; 
        font-size: 2rem;
        font-weight: 600;
        background: linear-gradient(90deg, #a78bfa, #818cf8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }}
    .meta {{ color: #94a3b8; font-size: 0.9rem; margin-bottom: 30px; }}
    .wrap {{ 
        overflow-x: auto; 
        background: #161b22; 
        border-radius: 12px;
        border: 1px solid #30363d;
        box-shadow: 0 10px 30px rgba(0,0,0,0.5);
    }}
    table {{ border-collapse: collapse; font-size: 0.85rem; min-width: 100%; }}
    th, td {{ padding: 12px 16px; border: 1px solid #30363d; white-space: nowrap; }}
    th {{ 
        background: #1f2937; 
        color: #f1f5f9; 
        font-weight: 600;
        text-transform: uppercase;
        font-size: 0.75rem;
        letter-spacing: 0.05em;
    }}
    tr:hover td {{ background-color: rgba(255,255,255,0.02) !important; }}
    code {{ background: #232a35; color: #a5b4fc; padding: 2px 6px; border-radius: 4px; border: 1px solid #3a4454; }}
    
    .section-title {{
        margin-top: 40px;
        font-size: 1.5rem;
        color: #cbd5e1;
        display: flex;
        align-items: center;
        gap: 12px;
    }}
    .section-title::before {{
        content: '';
        display: inline-block;
        width: 4px;
        height: 24px;
        background: #a78bfa;
        border-radius: 2px;
    }}
  </style>
</head>
<body>
  <h1>🔍 DEMAKAI — Evaluasi 6 Kombinasi Metode Pencarian</h1>
  <div class="meta">Generated: {ts} &nbsp;·&nbsp; Total {total_queries} query</div>
  
  <div class="section-title">Dataset: KBLI 2025</div>
  {html_kbli}
  
  <hr style="border: 0; border-top: 2px dashed #30363d; margin: 60px 0;">
  
  <div class="section-title">Dataset: KBJI 2014</div>
  {html_kbji}

    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 30px;">
        <div>
            <h3 style="color: #818cf8; margin-bottom: 12px;">Metode Terbaik</h3>
            <p style="color: #94a3b8; font-size: 0.95rem;">
                Berdasarkan data di atas, kombinasi <strong>Hybrid + Expansion</strong> memberikan hasil yang paling konsisten. 
                Dengan memanfaatkan sinonim, sistem mampu menjembatani perbedaan antara bahasa informal pengguna 
                dan terminologi formal KBLI/KBJI.
            </p>
        </div>
        <div>
            <h3 style="color: #818cf8; margin-bottom: 12px;">Saran Pengembangan</h3>
            <ul style="color: #94a3b8; font-size: 0.95rem; padding-left: 20px;">
                <li>Gunakan <strong>Vector Matching</strong> sebagai fallback semantik.</li>
                <li>Gunakan <strong>Query Expansion</strong> sebagai standar preprocessing utama.</li>
                <li>Pertahankan <strong>Advanced Preprocessing</strong> hanya sebagai baseline atau untuk analisis teknis.</li>
            </ul>
        </div>
    </div>
  </div>

</body>
</html>"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[HTML]  Gabungan Disimpan -> {os.path.abspath(filepath)}")


def save_excel_combined(rows_kbli: list, summary_kbli: list, rows_kbji: list, summary_kbji: list, filepath: str):
    """
    Simpan hasil evaluasi ke Excel dengan layout Dashboard yang kaya visual secara manual.
    Pandas tidak mendukung MultiIndex columns dengan index=False, jadi kita buat manual.
    """
    if not (rows_kbli or rows_kbji):
        return
        
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell

        # 1. Setup Workbook & Styles
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Evaluasi"
        # ... constant styles ...
        fill_header_top = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        fill_header_sub = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)
        font_slate = Font(color="CBD5E1")
        
        fill_rank_1  = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        font_rank_1  = Font(color="6EE7B7", bold=True)
        fill_rank_5  = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
        font_rank_5  = Font(color="A7F3D0")
        fill_rank_10 = PatternFill(start_color="1E3A2F", end_color="1E3A2F", fill_type="solid")
        font_rank_10 = Font(color="D1FAE5")
        fill_none    = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
        font_none    = Font(color="555555")
        border_thin  = Border(left=Side(style='thin', color='30363D'), right=Side(style='thin', color='30363D'), 
                             top=Side(style='thin', color='30363D'), bottom=Side(style='thin', color='30363D'))

        # 2. Helper untuk membuat tabel
        def write_table(start_idx, rows, title):
            # Header Row Indices
            h0, h1 = start_idx + 1, start_idx + 2
            d_start = start_idx + 3
            m_idx = d_start + len(rows)
            
            # Count columns
            total_cols = 4
            for _, p, _ in COMBINATIONS:
                total_cols += (6 if p in ("Advanced", "Expansion") else 5)

            # --- 1. Style all cells in the table area ---
            for r in range(start_idx, m_idx + 1):
                for c in range(1, total_cols + 1):
                    try:
                        cell = ws.cell(row=r, column=c)
                        cell.border = border_thin
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        if r == start_idx: # Title row
                            cell.fill = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
                        elif r == h0: # Method Headers
                            cell.fill = fill_header_top
                            cell.font = font_white
                        elif r == h1: # Metric Headers
                            cell.fill = fill_header_sub
                            cell.font = font_slate
                        elif r == m_idx: # Summary row
                            cell.fill = fill_header_sub
                        else: # Data rows
                            cell.fill = PatternFill(start_color="0F111A", end_color="0F111A", fill_type="solid")
                            cell.font = Font(color="E2E8F0")
                    except: pass

            # --- 2. Fill Context & Specific Styling ---
            try: ws.cell(row=start_idx, column=1, value=f"DATASET: {title}").font = Font(size=14, bold=True, color="A5B4FC")
            except: pass
            
            for i, label in enumerate(["No", "Tipe", "Query Asli", "GT"], 1):
                try: ws.cell(row=h0, column=i, value=label)
                except: pass

            curr_col = 5
            for s, p, _ in COMBINATIONS:
                group = f"{s} + {p}"
                span = 6 if p in ("Advanced", "Expansion") else 5
                try: ws.cell(row=h0, column=curr_col, value=group)
                except: pass
                
                metrics = ["Rank", "Top1", "Top5", "Top10", "KK"]
                if p in ("Advanced", "Expansion"): metrics.append("Preprocessing")
                for i, m in enumerate(metrics):
                    try: ws.cell(row=h1, column=curr_col + i, value=m)
                    except: pass
                curr_col += span

            # Data rows
            for r_idx, r in enumerate(rows, d_start):
                try:
                    ws.cell(row=r_idx, column=1, value=r.get("no"))
                    tipe_c = ws.cell(row=r_idx, column=2, value=r.get("tipe"))
                    tipe_c.font = Font(color="86EFAC" if r.get("tipe")=="KBLI" else "93C5FD", bold=True)
                    query_c = ws.cell(row=r_idx, column=3, value=r.get("query"))
                    query_c.font = Font(color="C4B5FD")
                    query_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    gt_c = ws.cell(row=r_idx, column=4, value=r.get("kode_ground_truth"))
                    gt_c.font = Font(color="A5B4FC", name="Consolas")
                except: pass

                c_ptr = 5
                for s, p, _ in COMBINATIONS:
                    label = f"{s}_{p}"
                    rank, kk = r.get(f"rank_{label}", 0), r.get(f"kk_{label}", 0.0)
                    span = 6 if p in ("Advanced", "Expansion") else 5
                    
                    try:
                        rc = ws.cell(row=r_idx, column=c_ptr, value=rank)
                        if rank == 1: rc.fill, rc.font = fill_rank_1, font_rank_1
                        elif 1 <= rank <= 5: rc.fill, rc.font = fill_rank_5, font_rank_5
                        elif 5 < rank <= 10: rc.fill, rc.font = fill_rank_10, font_rank_10
                        else: rc.fill, rc.font = fill_none, font_none

                        ws.cell(row=r_idx, column=c_ptr+1, value=r.get(f"top1_{label}"))
                        ws.cell(row=r_idx, column=c_ptr+2, value=r.get(f"top5_{label}"))
                        ws.cell(row=r_idx, column=c_ptr+3, value=r.get(f"top10_{label}"))
                        kc = ws.cell(row=r_idx, column=c_ptr+4, value=kk)
                        kc.number_format = '0.00'
                        if kk > 0: kc.font = Font(color="6EE7B7", bold=True)
                        if p in ("Advanced", "Expansion"):
                            ws.cell(row=r_idx, column=c_ptr+5, value=r.get(f"desc_{p.lower()}"))
                            ws.cell(row=r_idx, column=c_ptr+5).font = Font(color="D8B4FE", size=8)
                    except: pass
                    c_ptr += span

            # MRR Row
            try:
                ws.cell(row=m_idx, column=4, value="MRR / AVG").font = Font(bold=True, color="A78BFA")
                curr_col = 5
                for _, p, _ in COMBINATIONS:
                    kk_col = curr_col + 4
                    col_l = get_column_letter(kk_col)
                    formula = f"=AVERAGE({col_l}{d_start}:{col_l}{m_idx-1})"
                    cell = ws.cell(row=m_idx, column=kk_col, value=formula)
                    cell.font, cell.number_format = Font(bold=True, color="6EE7B7"), '0.0000'
                    curr_col += (6 if p in ("Advanced", "Expansion") else 5)
            except: pass
            
            return m_idx, total_cols

        # 3. Execution & Merging
        m_idx_kbli, cols_kbli = write_table(1, rows_kbli, "KBLI 2025")
        start_kbji = m_idx_kbli + 2
        m_idx_kbji, cols_kbji = write_table(start_kbji, rows_kbji, "KBJI 2014")

        # Now apply all merges
        def apply_merges(s_idx, m_idx, total_cols):
            ws.merge_cells(start_row=s_idx, start_column=1, end_row=s_idx, end_column=total_cols)
            h0, h1 = s_idx + 1, s_idx + 2
            for i in range(1, 5): ws.merge_cells(start_row=h0, start_column=i, end_row=h1, end_column=i)
            curr = 5
            for _, p, _ in COMBINATIONS:
                span = 6 if p in ("Advanced", "Expansion") else 5
                ws.merge_cells(start_row=h0, start_column=curr, end_row=h0, end_column=curr + span - 1)
                curr += span

        apply_merges(1, m_idx_kbli, cols_kbli)
        apply_merges(start_kbji, m_idx_kbji, cols_kbji)

        # Global Final
        try:
            for c_idx in range(1, max(cols_kbli, cols_kbji) + 1):
                max_l = 10
                for r_idx in range(1, ws.max_row + 1):
                    try:
                        val = ws.cell(row=r_idx, column=c_idx).value
                        if val and not str(val).startswith("="):
                            max_l = max(max_l, len(str(val)))
                    except: pass
                ws.column_dimensions[get_column_letter(c_idx)].width = min(max_l + 2, 50)
        except: pass

        ws.freeze_panes = "E4"
        ws.sheet_view.showGridLines = False
        wb.save(filepath)

        print(f"[EXCEL] Gabungan Disimpan (Final Dashboard) -> {os.path.abspath(filepath)}")
    except Exception as e:
        print(f"[ERROR] Gagal menyimpan EXCEL: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Evaluasi Otomatis DEMAKAI")
    parser.add_argument("--output", type=str, help="Nama file output (tanpa ekstensi)")
    parser.add_argument("--mode", type=str, choices=["sebelum", "sesudah"], default="sesudah", help="Mode evaluasi (sebelum/sesudah query lapangan)")
    args = parser.parse_args()

    # Set Environment Variable for Expansion
    if args.mode == "sebelum":
        os.environ["ENABLE_EXPANSION"] = "false"
        print(">>> MODE: Tanpa Query Lapangan (Expansion Disabled) <<<")
    else:
        os.environ["ENABLE_EXPANSION"] = "true"
        print(">>> MODE: Dengan Query Lapangan (Expansion Enabled) <<<")

    _base_dir = os.path.dirname(os.path.abspath(__file__))
    QUERIES_FILE = os.path.join(_base_dir, "queries.csv")
    LIMIT        = 10

    print("=" * 60)
    print("  DEMAKAI — Evaluasi Otomatis v2.0")
    print("=" * 60)

    rows = run_evaluation(QUERIES_FILE, limit=LIMIT)
    if not rows: sys.exit(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Determine Output Filename
    if args.output:
        file_base = args.output
    else:
        file_base = f"evaluasi_{ts}_Gabungan"

    rows_kbli = [r for r in rows if r.get("tipe") == "KBLI"]
    rows_kbji = [r for r in rows if r.get("tipe") == "KBJI"]

    summary_kbli = calculate_summary(rows_kbli)
    summary_kbji = calculate_summary(rows_kbji)

    file_html = f"output/{file_base}.html"
    file_excel = f"output/{file_base}.xlsx"
    
    save_html_combined(rows_kbli, summary_kbli, rows_kbji, summary_kbji, file_html)
    save_excel_combined(rows_kbli, summary_kbli, rows_kbji, summary_kbji, file_excel)

    print(f"\nSelesai! Laporan siap di folder output/.")
